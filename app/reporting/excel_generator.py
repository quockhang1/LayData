import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import sqlite3
import json
import os
from app.database import get_connection

def generate_excel_report(output_dir: str = "."):
    """
    Generate professional Excel workbook with:
    Sheet 1: Raw Products
    Sheet 2: Grouped Products
    Sheet 3: Price Comparison (with colors)
    Sheet 4: Top Best Deals
    Sheet 5: Top Highest Difference
    Sheet 6: Failed URLs
    Sheet 7: Statistics
    """
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, "products_compare.xlsx")
    
    conn = get_connection()
    
    # Read tables into Pandas dataframes
    raw_df = pd.read_sql_query("SELECT * FROM products_raw", conn)
    grouped_df = pd.read_sql_query("SELECT * FROM products_grouped", conn)
    
    # Comparison Query
    comparison_query = """
    SELECT 
        pg.canonical_name as product_name,
        pg.model,
        MIN(pr.final_price) as lowest_price,
        MAX(pr.final_price) as highest_price,
        (MAX(pr.final_price) - MIN(pr.final_price)) as price_difference,
        ROUND(((MAX(pr.final_price) - MIN(pr.final_price)) / MIN(pr.final_price)) * 100, 2) as difference_percent,
        COUNT(DISTINCT pr.website) as store_count
    FROM products_grouped pg
    JOIN product_group_mapping pgm ON pg.id = pgm.grouped_id
    JOIN products_raw pr ON pgm.raw_id = pr.id
    GROUP BY pg.id
    """
    compare_df = pd.read_sql_query(comparison_query, conn)
    
    # Failed URLs
    failed_df = pd.read_sql_query("SELECT * FROM failed_urls", conn)
    
    # Build Excel
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Colors for highlighting
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light Green for Lowest
    fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")    # Light Red for Highest
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light Yellow for Large Diff
    fill_orange = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid") # Orange for missing
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Helper: write sheets
    def write_sheet(df, title, is_comparison=False):
        ws = wb.create_sheet(title=title)
        
        # Write headers
        headers = list(df.columns)
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        # Write data rows
        for _, row in df.iterrows():
            row_vals = [row[col] for col in df.columns]
            ws.append(row_vals)
            
        # Formatting rows
        for r_idx in range(2, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.alignment = align_left
                
                # Highlight missing data as Orange (empty or zero values for crucial product keys)
                col_name = headers[c_idx - 1]
                if col_name in ["model", "brand", "sku", "barcode", "image", "description", "category", "price", "sale_price"]:
                    if cell.value is None or str(cell.value).strip() == "" or str(cell.value).lower() == "nan" or cell.value == 0 or cell.value == 0.0:
                        cell.fill = fill_orange
                        
                # Apply color rules to Price Comparison sheet
                if is_comparison:
                    if col_name == "lowest_price":
                        cell.fill = fill_green
                    elif col_name == "highest_price":
                        cell.fill = fill_red
                    elif col_name == "price_difference" and cell.value and float(cell.value) > 2000000:
                        cell.fill = fill_yellow
                        
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    # Write Sheet 1: Raw Products
    write_sheet(raw_df, "Raw Products")
    
    # Write Sheet 2: Grouped Products
    write_sheet(grouped_df, "Grouped Products")
    
    # Write Sheet 3: Price Comparison
    write_sheet(compare_df, "Price Comparison", is_comparison=True)
    
    # Write Sheet 4: Top Best Deals (Groups with lowest price offers)
    best_deals_query = """
    SELECT 
        pg.canonical_name as product_name,
        pr.website as store,
        pr.final_price as best_price,
        pr.url as product_url
    FROM products_grouped pg
    JOIN product_group_mapping pgm ON pg.id = pgm.grouped_id
    JOIN products_raw pr ON pgm.raw_id = pr.id
    WHERE pr.final_price = (
        SELECT MIN(final_price) 
        FROM products_raw 
        WHERE id IN (SELECT raw_id FROM product_group_mapping WHERE grouped_id = pg.id)
    )
    """
    best_deals_df = pd.read_sql_query(best_deals_query, conn)
    write_sheet(best_deals_df, "Top Best Deals")
    
    # Write Sheet 5: Top Highest Difference
    highest_diff_df = compare_df.sort_values(by="price_difference", ascending=False).head(20)
    write_sheet(highest_diff_df, "Top Highest Difference", is_comparison=True)
    
    # Write Sheet 6: Failed URLs
    write_sheet(failed_df, "Failed URLs")
    
    # Write Sheet 7: Statistics
    ws_stats = wb.create_sheet(title="Statistics")
    stats_data = [
        ["Metric", "Value"],
        ["Total URLs", conn.execute("SELECT COUNT(*) FROM urls").fetchone()[0]],
        ["Successful URLs", conn.execute("SELECT COUNT(*) FROM urls WHERE status = 'completed'").fetchone()[0]],
        ["Failed URLs", conn.execute("SELECT COUNT(*) FROM urls WHERE status = 'failed'").fetchone()[0]],
        ["Total Products Found", conn.execute("SELECT COUNT(*) FROM products_raw").fetchone()[0]],
        ["Total Product Groups", conn.execute("SELECT COUNT(*) FROM products_grouped").fetchone()[0]],
        ["Average Price (VND)", conn.execute("SELECT ROUND(AVG(final_price), 2) FROM products_raw").fetchone()[0] or 0.0]
    ]
    for row in stats_data:
        ws_stats.append(row)
        
    for r_idx in range(1, len(stats_data) + 1):
        for c_idx in range(1, 3):
            cell = ws_stats.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 15
    
    wb.save(file_path)
    wb.close()
    conn.close()
    return file_path
